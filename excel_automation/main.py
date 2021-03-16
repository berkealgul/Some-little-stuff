import openpyxl
from openpyxl.styles import PatternFill


def main(actionCode, input_name):
    wb = openpyxl.load_workbook(input_name+".xlsx")
    ws = wb.active

    done = True

    if actionCode == "0":
        step1(wb, input_name)
        ws = step2(ws)
        ws = step3(ws)
        ws = step4(ws)
        ws = step5(ws)
        ws = step6(ws)
        ws = step7(ws)
        ws = step8(ws)
        ws = step9(ws)
    elif actionCode == "1":
        step1(wb, input_name)
    elif actionCode == "2":
        ws = step2(ws)
    elif actionCode == "3":
        ws = step3(ws)
    elif actionCode == "4":
        ws = step4(ws)
    elif actionCode == "5":
        ws = step5(ws)
    elif actionCode == "6":
        ws = step6(ws)
    elif actionCode == "7":
        ws = step7(ws)
    elif actionCode == "8":
        ws = step8(ws)
    elif actionCode == "9":
        ws = step9(ws)
    else:
        print("UNVALID action code!!! Try again")
        done = False

    if done is True:
        print("Done!")

    return wb, done


def greet():
    print("Hello user")
    print("BEFORE use this software consider following assumptions\n")
    print("1-input file must be .xlsx format and must be stored in same folder with code")
    print("2-output files will be in xlsx format and will be stored in same folder with program")
    print("3-Before modify input file you must close other programs that monitors our excel file such as Ms Excel"
          " otherwise it will give -permission denied- error")
    print("if you perform step 1 the result will be saved as \"name you entered\"-copy.xlsx")

    name = str(input("Enter the name of input file (dont put .xlsx): "))
    return name


def sayGoodbye():
    print("\nThanks for using software")
    print("Goodbye :(")


def interface():
    actionCode = input("\nEnter step code you want to perform from 1 to 9 (if you "
                        "want to Run All type 0, 10 to exit): ")
    return actionCode


def save_result(wb, name):
    done = False
    while done is False:
        try:
            done = True
            wb.save(name+".xlsx")
        except:
            input("\nFAILED!! It probably gave permission deny error. Close excel programs and enter anything to try again")
            done = False


def step1(wb, name):
    wb.save(name+"-copy.xlsx")
    print("Step 1 done")


def step2(ws):
    rows = ws.max_row
    i = 2
    while i <= rows:
        phoneNum = ws['G'+str(i)].value
        if isEmpty(phoneNum):
            ws.delete_rows(i, 1)
            rows -= 1
        else:
            i += 1

    print("Step 2 done")
    return ws


def step3(ws):
    rows = ws.max_row
    i = 2
    while i <= rows:
        phoneNum = ws['G' + str(i)].value
        deleteRow = False

        if isEmpty(phoneNum):
            deleteRow = False
        elif count_digits(str(phoneNum)) < 7:
            deleteRow = True

        if deleteRow is True:
            ws.delete_rows(i, 1)
            rows -= 1
        else:
            i += 1

    print("Step 3 done")
    return ws


def step4(ws):
    rows = ws.max_row
    i = 2
    while i <= rows:
        text = ws['B' + str(i)].value
        if searchForWord(text, "suche") or searchForWord(text, "SUCHE") or searchForWord(text, "Suche"):
            ws.delete_rows(i, 1)
            rows -= 1
        else:
            i += 1
    print("Step 4 done")
    return ws


def step5(ws):
    queried = input("Enter queried word for step 5: ")
    ws['M1'].value = "Region"
    ws['M1'].fill = PatternFill(start_color='f7f56f',
                   end_color='f7f56f',
                   fill_type='solid')
    for i in range(2, ws.max_row+1):
        ws['M' + str(i)].value = str(queried)
    print("Step 5 done")
    return ws


def step6(ws):
    for i in range(2, ws.max_row+1):
        text = ws['J' + str(i)].value
        if text.find("Außenstellplatz") != -1:
            ws['J' + str(i)].value = "Stellplatz"
        i += 1
    print("Step 6 done")
    return ws


def step7(ws):
    for i in range(2, ws.max_row+1):
        ws['K' + str(i)].value = None
    print("Step 7 done")
    return ws


def step8(ws):
    word_list = get_word_list()
    for i in range(2, ws.max_row+1):
        found = ""
        for word in word_list:
            if ws['L' + str(i)].value.find(word) != -1:
                found += word + ", "
        ws['K' + str(i)].value = found[:-2]
    print("Step 8 done")
    return ws


def step9(ws):
    for i in range(2, ws.max_row+1):
        if ws['C'+str(i)].value.find("VB") != -1:
            ws['L'+str(i)].value += '\n\nDer Preis ist Verhandlungsbasis'
    print("Step 9 done")
    return ws


def count_digits(input_str):
    digit_c = 0
    for c in input_str:
        if c.isnumeric():
            digit_c += 1
    return digit_c


def get_word_list():
    word_list = list()
    while True:
        g = input("Enter a word for word-list for step 8 (enter 0 for exit): ")
        g = str(g)
        if g == "0":
            break
        word_list.append(g)
    return word_list


def isEmpty(data):
    if data is None:
        return True
    else:
        if str(data).replace(" ", "") == "":
            return True
    return False


# If this is true row should be deleted
def searchForWord(Str, target):
    idx = Str.find(target)
    if idx == -1:
        return False

    frontEmpty = False
    backEmpty = False

    try:
        if Str[idx + len(target)] == " ":
            frontEmpty = True
    except IndexError:
        frontEmpty = True

    try:
        if Str[idx - 1] == " " or Str[idx - 1] == Str[-1]:
            backEmpty = True
    except IndexError:
        backEmpty = True

    if frontEmpty and backEmpty is True:
        return True
    else:
        return False


if __name__ == "__main__":
    input_name = greet()
    while True:
        actionCode = interface()
        if actionCode == "10":
            break
        wb, done = main(actionCode, input_name)

        if actionCode != "1" and done is True:
            g = input("\nActions done press s to save the result, press other for not save: ")
            if str(g) == "s":
                save_result(wb, input_name)
                print("Saved!")
            else:
                print("Not saved!!")
    sayGoodbye()
